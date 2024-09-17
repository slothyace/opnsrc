# Python Standard Library
import os
import sys
import subprocess
import csv
import glob
import tkinter as tk
from tkinter import filedialog, messagebox

# External Libraries
installables = ["ttkbootstrap", "openpyxl"]
for library in installables:
    try:
        __import__(library)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", library])

import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import openpyxl

# Log errors to a file
def log_error(error_message):
    with open("error_log.txt", "a") as log_file:
        log_file.write(error_message + "\n")

# Check if the script is running with pythonw
if not sys.executable.endswith('pythonw.exe'):
    try:
        os.execv(sys.executable.replace('python.exe', 'pythonw.exe'), ['pythonw.exe'] + sys.argv)
    except Exception as e:
        log_error(f"Failed to relaunch with pythonw: {e}")
else:
    try:
        # Function to delete file if user chooses to do so
        def delFile(console, file):
            delete_choice = messagebox.askyesno("Delete File", f"Do you want to delete the original file: {file}?")
            if delete_choice:
                try:
                    os.remove(file)
                    console.insert(END, f"Deleted {file}\n\n")
                except Exception as e:
                    console.insert(END, f"Failed to delete {file} due to: {e}\n\n")
            else:
                console.insert(END, f"{file} was not deleted.\n\n")
            console.see(END)
            console.update_idletasks()

        # Function to ask once whether to delete all original files in a folder
        def delFile_Folder(console, files):
            delete_choice = messagebox.askyesno("Delete Files", "Do you want to delete the original files?")
            if delete_choice:
                for file in files:
                    try:
                        os.remove(file)
                        console.insert(END, f"Deleted {file}\n\n")
                    except Exception as e:
                        console.insert(END, f"Failed to delete {file} due to: {e}\n\n")
                    console.see(END)
                    console.update_idletasks()
            else:
                console.insert(END, "Files were not deleted.\n\n")
                console.see(END)
                console.update_idletasks()

        # Function to convert CSV to XLSX
        def C2X(file_path, console):
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                with open(file_path, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    for row in reader:
                        ws.append(row)
                xlsx_file = file_path.replace('.csv', '.xlsx')
                wb.save(xlsx_file)
                console.insert(END, f"Converted {file_path} to {xlsx_file}\n\n")
                console.see(END)
                console.update_idletasks()
                delFile(console, file_path)  # Ask to delete the original file
            except Exception as e:
                console.insert(END, f"Failed to convert {file_path} due to: {e}\n\n")
                console.see(END)
                console.update_idletasks()

        # Function to convert XLSX to CSV
        def X2C(file_path, console):
            try:
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                csv_file = file_path.replace('.xlsx', '.csv')
                with open(csv_file, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    for row in ws.iter_rows(values_only=True):
                        writer.writerow(row)
                console.insert(END, f"Converted {file_path} to {csv_file}\n\n")
                console.see(END)
                console.update_idletasks()
                delFile(console, file_path)  # Ask to delete the original file
            except Exception as e:
                console.insert(END, f"Failed to convert {file_path} due to: {e}\n\n")
                console.see(END)
                console.update_idletasks()

        # Function to convert all CSV files in a folder to XLSX
        def C2X_Folder(folder_path, console):
            csv_files = glob.glob(os.path.join(folder_path, '*.csv'))
            if not csv_files:
                console.insert(END, f"No CSV files found in the folder '{folder_path}'.\n\n")
                console.see(END)
                console.update_idletasks()
                return
            converted_files = []
            for csv_file in csv_files:
                C2X(csv_file, console)
                converted_files.append(csv_file)  # Track only original CSV files
            delFile_Folder(console, converted_files)  # Ask to delete only the original files

        # Function to convert all XLSX files in a folder to CSV
        def X2C_Folder(folder_path, console):
            xlsx_files = glob.glob(os.path.join(folder_path, '*.xlsx'))
            if not xlsx_files:
                console.insert(END, f"No XLSX files found in the folder '{folder_path}'.\n\n")
                console.see(END)
                console.update_idletasks()
                return
            converted_files = []
            for xlsx_file in xlsx_files:
                X2C(xlsx_file, console)
                converted_files.append(xlsx_file)  # Track only original XLSX files
            delFile_Folder(console, converted_files)  # Ask to delete only the original files

        # GUI Functions for browsing and folder/file selection
        def browse_C2X_Folder(console):
            folder_path = filedialog.askdirectory()
            if folder_path:
                C2X_Folder(folder_path, console)

        def browse_X2C_Folder(console):
            folder_path = filedialog.askdirectory()
            if folder_path:
                X2C_Folder(folder_path, console)

        def sC2X(console):
            file_path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
            if file_path:
                C2X(file_path, console)

        def sX2C(console):
            file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
            if file_path:
                X2C(file_path, console)

        # GUI Part
        root = ttk.Window(themename="darkly")
        root.title("Acedia Slothy Tools")
        root.geometry("500x300")
        root.resizable(True, True)
        root.minsize(500, 300)

        button_frame = ttk.Frame(root)
        button_frame.pack(pady=8, padx=8)

        button_csv_to_xlsx_folder = ttk.Button(button_frame, text="CSV -> XLSX (Folder)", command=lambda: browse_C2X_Folder(console), bootstyle="success-outline", width=20)
        button_csv_to_xlsx_folder.grid(row=0, column=0, padx=2, pady=2)

        button_xlsx_to_csv_folder = ttk.Button(button_frame, text="XLSX -> CSV (Folder)", command=lambda: browse_X2C_Folder(console), bootstyle="success-outline", width=20)
        button_xlsx_to_csv_folder.grid(row=1, column=0, padx=2, pady=2)

        button_csv_to_xlsx_file = ttk.Button(button_frame, text="CSV -> XLSX (File)", command=lambda: sC2X(console), bootstyle="success-outline", width=20)
        button_csv_to_xlsx_file.grid(row=0, column=1, padx=2, pady=2)

        button_xlsx_to_csv_file = ttk.Button(button_frame, text="XLSX -> CSV (File)", command=lambda: sX2C(console), bootstyle="success-outline", width=20)
        button_xlsx_to_csv_file.grid(row=1, column=1, padx=2, pady=2)

        separator = ttk.Separator(root, orient='horizontal')
        separator.pack(fill='x', pady=0)

        console = ttk.Text(root, wrap=tk.WORD, width=600, height=15)
        console.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)
        console.bind("<Key>", lambda e: "break")

        root.mainloop()
    except Exception as e:
        log_error(f"Failed to initialize GUI: {e}")
